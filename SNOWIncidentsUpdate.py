import os
import sys
import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

# WorkingDirectory = os.getcwd()
# print(getpass.getuser())
# print(WorkingDirectory)
# ExcelFileLoc=WorkingDirectory+"\\Incidents.xlsx"
wb = load_workbook("Incidents.xlsx")
sheet = wb["Sheet1"]
row_count = sheet.max_row
column_count = sheet.max_column
print("No of Incidents is " + str(row_count-1))

print("Opening Chrome...")
# PATH = "C:\Program Files (x86)\chromedriver.exe"
if getattr(sys, 'frozen', False):
    # executed as a bundled exe, the driver is in the extracted folder
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    # executed as a simple script, the driver should be in `PATH`
    # driver = webdriver.Chrome(ChromeDriverManager().install())
    driver = webdriver.Chrome()
    driver.maximize_window()
# driver.set_page_load_timeout(30)

print("Opening servicenow webpage...")
driver.get("https://gaptech.service-now.com/navpage.do")
wait = WebDriverWait(driver, 10)

for x in range(2,row_count+1):
    print("Progress : " + str(x-1) + " of " + str(row_count-1))
    statusUpdate = sheet["L" + str(x)]
    incident_no = sheet["B" + str(x)].value
    if statusUpdate.value == 'Updated':
        print(incident_no + "already updated as per sheet")
        continue
    driver.switch_to.default_content()
    search = wait.until(EC.presence_of_element_located((By.ID, "sysparm_search")))
    search.clear()
    print("Opening Incident no : " + incident_no)
    search.send_keys(incident_no)
    time.sleep(2)
    search.send_keys(Keys.RETURN)
    try:
        WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                        'Timed out waiting for update ' +
                                        'confirmation popup to appear.')
        alert = driver.switch_to.alert
        print("\tAlert while opening next incident" + alert.text + ". Accepting and continuing")
        alert.accept()
    except TimeoutException:
        print("No popup")
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID,'gsft_main')))
    time.sleep(2)
    try:
        IncidentNoLoadOut = wait.until(EC.presence_of_element_located((By.ID, "sys_readonly.incident.number")))
        IncidentState = Select(driver.find_element_by_id("sys_readonly.incident.state"))
        incidentSelectedStateText = IncidentState.first_selected_option.text
        if incidentSelectedStateText == "Resolved" or incidentSelectedStateText == "Closed":
            print("\t" + incident_no + " Already resolved/closed as per SNOW. Continuing with other incidents!!")
            statusUpdate.value = "Updated"
            wb.save("Incidents.xlsx")
            continue
    except NoSuchElementException:
        assignmentGroupValidation = sheet["D" + str(x)].value
        assignmentGroup = wait.until(EC.presence_of_element_located((By.ID, "sys_display.incident.assignment_group"))).get_attribute('value')
        if  assignmentGroup!= assignmentGroupValidation:
            print("\tThe incident " + incident_no + " not assigned to group " + assignmentGroupValidation)
            statusUpdate.value = "Incident Not assigned to appropriate group "
            continue
        try:
            IncidentState = Select(wait.until(EC.presence_of_element_located((By.ID, "incident.state"))))
            IncidentState.select_by_value('6')
            print("\tClicked on resolved..")
            time.sleep(3)
        except TimeoutException:
            print("\tError opening incident : " + incident_no + " Please check!")


#Select brd value
    print("\tSelecting Gap as brand value..")
    brand = wait.until(EC.presence_of_element_located((By.ID, "incident.u_brand_edit")))
    driver.execute_script("arguments[0].removeAttribute('style')", brand)
    time.sleep(2)
    wait.until(EC.presence_of_element_located((By.ID,'sys_display.incident.u_brand'))).send_keys("Gap")
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@id='AC.incident.u_brand']/div/span[text()='Gap']"))).click()
    except TimeoutException:
        main_window = driver.current_window_handle
        # wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'gsft_main')))
        brd_Button_Click = wait.until(EC.presence_of_element_located((By.ID, 'lookup.incident.u_brand')))
        brd_Button_Click.send_keys(Keys.ENTER)
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) == 2)
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        bvalue = wait.until(EC.presence_of_element_located((By.XPATH,"//div[@class='input-group']//input[@placeholder='Search']")))
        bvalue.clear()
        bvalue.send_keys("Gap")
        bvalue.send_keys(Keys.ENTER)
        wait.until(EC.presence_of_element_located((By.LINK_TEXT,'Gap'))).click()
        driver.switch_to.window(main_window)
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID,'gsft_main')))

#Select market value
    print("\tSelecting Global as market value")
    market = wait.until(EC.presence_of_element_located((By.ID,"incident.u_market_edit")))
    driver.execute_script("arguments[0].removeAttribute('style')", market)
    time.sleep(2)
    wait.until(EC.presence_of_element_located((By.ID, 'sys_display.incident.u_market'))).send_keys("Global")
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='AC.incident.u_market']/div/span[text()='Global']"))).click()
    except TimeoutException:
        main_window = driver.current_window_handle
        mkt_Button_Click = wait.until(EC.element_to_be_clickable((By.ID, 'lookup.incident.u_market')))
        mkt_Button_Click.send_keys(Keys.ENTER)
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) == 2)
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        mvalue = wait.until(EC.presence_of_element_located((By.XPATH,"//div[@class='input-group']//input[@placeholder='Search']")))
        mvalue.clear()
        mvalue.send_keys("global")
        mvalue.send_keys(Keys.ENTER)
        wait.until(EC.presence_of_element_located((By.LINK_TEXT,'Global'))).click()
        driver.switch_to.window(main_window)
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID,'gsft_main')))

#Enter Assigned to name
    assignedToName = sheet["E"+str(x)].value
    print("\tAssigning incident to " + assignedToName)
    assignedTo = wait.until(EC.presence_of_element_located((By.ID,'sys_display.incident.assigned_to')))
    assignedTo.clear()
    assignedTo.send_keys(assignedToName)
    time.sleep(2)
    assignedTo.send_keys(Keys.RETURN)
#Enter AppCI details
    print("\tSelecting AppCI details..")
    appInfra = Select(driver.find_element_by_id("incident.u_app_infra"))
    appInfra.select_by_value('app')
#Enter Response type
    print("\tSelecting response type...")
    responseType = Select(driver.find_element_by_id("incident.u_response_type"))
    responseType.select_by_value('Reactive')
#Toggel closure information
    closureInfo = wait.until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Closure Information')]")))
    driver.execute_script("arguments[0].scrollIntoView();", closureInfo)
    closureInfo.click()
#Enter closure code
    print("\tSelecting solved by workaround")
    closureCode = Select(wait.until(EC.presence_of_element_located((By.ID,"incident.close_code"))))
    closureCode.select_by_value('Solved by Workaround')
#Enter root cause code
    print("\tSelecting root cause code..")
    rootCauseCode = Select(wait.until(EC.presence_of_element_located((By.ID,"incident.u_root_cause_code"))))
    rootCauseCode.select_by_value('Application software failure')
    time.sleep(2)

    print("\tSelecting Sub closure code..")
    rmClosureCodeSub = wait.until(EC.presence_of_element_located((By.ID,"element.incident.u_closure_code_sub")))
    driver.execute_script("arguments[0].removeAttribute('style')", rmClosureCodeSub)
    time.sleep(2)
    closureCodeSub= Select(WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID,"incident.u_closure_code_sub"))))
    closureCodeSub.select_by_value('other')

    RootCauseNotesValue=sheet["G"+str(x)].value
    print("\tRoot cause notes : " + RootCauseNotesValue)
    RootCauseNotes=driver.find_element_by_id("incident.u_root_cause_notes")
    RootCauseNotes.clear()
    RootCauseNotes.send_keys(RootCauseNotesValue)

    ClosureNotesValue=sheet["H"+str(x)].value
    print("\tClosure notes : " + ClosureNotesValue)
    ClosureNotes=driver.find_element_by_id("incident.close_notes")
    ClosureNotes.clear()
    ClosureNotes.send_keys(ClosureNotesValue)
    print("\tSubmitting Incdident.. Please wait!")
    UpdateButton = wait.until(EC.element_to_be_clickable((By.ID,'sysverb_update')))
    UpdateButton.click()
    time.sleep(5)

    try:
        WebDriverWait(driver, 3).until(EC.alert_is_present(),
                                        'Timed out waiting for update ' +
                                        'confirmation popup to appear.')
        alert = driver.switch_to.alert
        alertReason = alert.text
        print("\tUnable to update incident due to " + str(alertReason) + ". Accepting the popup and continuing with next incidents")
        statusUpdate.value = alertReason
        alert.accept()
        continue
    except TimeoutException:
        print("\tConfirming incident update..")
        print("\t" + incident_no + " Successfully Updated")
        statusUpdate.value = 'Updated'
        wb.save("Incidents.xlsx")

    # try:
    #     if x !=2:
    #         wait.until(EC.presence_of_element_located(
    #             (By.XPATH,"// div[ @class ='outputmsg outputmsg_info notification notification-info']")))
    #
    #
    #     else:
    #         print(incident_no + " Successfully Updated")
    #         statusUpdate.value = 'Updated'
    #         wb.save("Incidents.xlsx")
    # except TimeoutException:
    #     print(incident_no + " Update failed")

print("Completed updating all incidents. Closing the browser now!")
driver.quit()